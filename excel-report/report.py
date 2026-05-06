"""CSV → branded XLSX report with charts, conditional formatting, and a pivot summary.

Demonstrates the kind of Excel automation I deliver for clients who have
raw export data and want a polished, share-ready workbook.

Usage:
    python report.py --input sample/sales.csv --output report.xlsx
    python report.py --input data.csv --output q1.xlsx --title "Q1 2026 Sales"
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# Brand palette — easy to swap for the client's colors
BRAND_PRIMARY = "2C5F8D"   # deep blue
BRAND_ACCENT = "F2A93C"    # orange
BRAND_LIGHT = "E8F0F7"     # light blue tint

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def load_data(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    # Try to coerce a date column
    for col in df.columns:
        if col.lower() in {"date", "order_date", "day"}:
            df[col] = pd.to_datetime(df[col], errors="coerce")
            break
    return df


def style_header(ws, row: int, ncols: int) -> None:
    fill = PatternFill("solid", fgColor=BRAND_PRIMARY)
    font = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def add_summary_sheet(wb: Workbook, df: pd.DataFrame, title: str) -> None:
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color=BRAND_PRIMARY)
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="888888")

    # KPI block
    ws["A4"] = "Key metrics"
    ws["A4"].font = Font(bold=True, size=12)

    kpis = []
    if "revenue" in df.columns:
        kpis.append(("Total revenue", f"{df['revenue'].sum():,.2f}"))
        kpis.append(("Average order value", f"{df['revenue'].mean():,.2f}"))
    if "orders" in df.columns:
        kpis.append(("Total orders", int(df['orders'].sum())))
    if "region" in df.columns:
        kpis.append(("Regions covered", df['region'].nunique()))
    kpis.append(("Rows in dataset", len(df)))

    for i, (label, value) in enumerate(kpis, start=5):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=BRAND_LIGHT)


def add_data_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Data")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    style_header(ws, 1, len(df.columns))

    # Conditional formatting on numeric columns
    for col_idx, col_name in enumerate(df.columns, start=1):
        if pd.api.types.is_numeric_dtype(df[col_name]):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            rng = f"{col_letter}2:{col_letter}{ws.max_row}"
            ws.conditional_formatting.add(
                rng,
                ColorScaleRule(
                    start_type="min", start_color="FFFFFF",
                    mid_type="percentile", mid_value=50, mid_color=BRAND_LIGHT,
                    end_type="max", end_color=BRAND_ACCENT,
                ),
            )

    autosize_columns(ws)
    ws.freeze_panes = "A2"


def add_pivot_chart(wb: Workbook, df: pd.DataFrame) -> None:
    """If we have region + revenue, build a bar chart of revenue per region."""
    if not {"region", "revenue"}.issubset(df.columns):
        return

    pivot = df.groupby("region", as_index=False)["revenue"].sum().sort_values("revenue", ascending=False)
    ws = wb.create_sheet("By region")
    for r in dataframe_to_rows(pivot, index=False, header=True):
        ws.append(r)
    style_header(ws, 1, len(pivot.columns))
    autosize_columns(ws)

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Revenue by region"
    chart.y_axis.title = "Region"
    chart.x_axis.title = "Revenue"
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20
    ws.add_chart(chart, "D2")


def add_timeseries_chart(wb: Workbook, df: pd.DataFrame) -> None:
    date_col = next((c for c in df.columns if pd.api.types.is_datetime64_any_dtype(df[c])), None)
    if not date_col or "revenue" not in df.columns:
        return

    daily = (
        df.dropna(subset=[date_col])
          .groupby(df[date_col].dt.date, as_index=False)["revenue"]
          .sum()
    )
    daily.columns = ["date", "revenue"]
    ws = wb.create_sheet("Daily revenue")
    for r in dataframe_to_rows(daily, index=False, header=True):
        ws.append(r)
    style_header(ws, 1, len(daily.columns))
    autosize_columns(ws)

    chart = LineChart()
    chart.title = "Daily revenue"
    chart.y_axis.title = "Revenue"
    chart.x_axis.title = "Date"
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20
    ws.add_chart(chart, "D2")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--title", default="Sales report")
    args = ap.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        print(f"Input not found: {in_path}", file=sys.stderr)
        return 1

    df = load_data(in_path)
    print(f"Loaded {len(df)} rows, {len(df.columns)} columns")

    wb = Workbook()
    # Drop the default sheet so our Summary becomes index 0
    default = wb.active
    wb.remove(default)

    add_summary_sheet(wb, df, args.title)
    add_data_sheet(wb, df)
    add_pivot_chart(wb, df)
    add_timeseries_chart(wb, df)

    out = Path(args.output)
    wb.save(out)
    print(f"Wrote {out} ({out.stat().st_size:,} bytes, {len(wb.sheetnames)} sheets)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
